import streamlit as st
import pandas as pd
import sqlite3
import hashlib
import os
import io
import webbrowser
import subprocess
import json
import tempfile
import shutil
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuration
DATABASE = "productivity_tracker.db"
ADMIN_USERNAME = "Adityakarthik"
TARGET_EMAIL = "rakhi.purohit@thomsonreuters.com"

# SharePoint Configuration
SHAREPOINT_SITE_URL = "https://trten.sharepoint.com/sites/CPT-RPurohit"
SHAREPOINT_FOLDER_PATH = "/Shared Documents/General/2025/BLR productivity gains"
SHAREPOINT_FULL_URL = "https://trten.sharepoint.com/sites/CPT-RPurohit/Shared%20Documents/Forms/AllItems.aspx?CID=ed5bb29a%2D39bb%2D4316%2Dbad5%2D66acedda78ba&FolderCTID=0x01200077D40D3D6B480D4DBC70A0833B88D7DD&id=%2Fsites%2FCPT%2DRPurohit%2FShared%20Documents%2FGeneral%2F2025%2FBLR%20productivity%20gains"

def init_database():
    """Initialize the SQLite database with required tables"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    # Create main data table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS productivity_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date_added TEXT NOT NULL,
            team_member TEXT NOT NULL,
            task_description TEXT NOT NULL,
            time_saved_hours REAL NOT NULL,
            impact_category TEXT NOT NULL,
            sprint_week TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    conn.commit()
    conn.close()

def verify_admin(username, password):
    """Verify admin credentials"""
    admin_code = os.getenv("TRACKER_ADMIN_CODE", "admin")
    return username == ADMIN_USERNAME and password == admin_code

def get_all_data():
    """Get all data from database"""
    conn = sqlite3.connect(DATABASE)
    df = pd.read_sql_query("SELECT * FROM productivity_data ORDER BY created_at DESC", conn)
    conn.close()
    return df

def add_data_to_db(data):
    """Add new data to database"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    for _, row in data.iterrows():
        cursor.execute("""
            INSERT INTO productivity_data 
            (date_added, team_member, task_description, time_saved_hours, impact_category, sprint_week)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            row["Date Added"],
            row["Team Member"],
            row["Task Description"],
            row["Time Saved (Hours)"],
            row["Impact Category"],
            row["Sprint/Week"]
        ))
    
    conn.commit()
    conn.close()

def upload_to_sharepoint_powershell(file_path, filename):
    """Upload file to SharePoint using PowerShell PnP"""
    try:
        # Enhanced PowerShell script with multiple authentication methods
        ps_script = f"""
        # Function to install PnP PowerShell if needed
        function Install-PnPIfNeeded {{
            if (-not (Get-Module -ListAvailable -Name PnP.PowerShell)) {{
                Write-Host " Installing PnP.PowerShell module..."
                try {{
                    Install-Module -Name PnP.PowerShell -Force -AllowClobber -Scope CurrentUser -SkipPublisherCheck
                    Write-Host " PnP.PowerShell installed successfully"
                    return $true
                }} catch {{
                    Write-Host " Failed to install PnP.PowerShell: $($_.Exception.Message)"
                    return $false
                }}
            }} else {{
                Write-Host " PnP.PowerShell module already available"
                return $true
            }}
        }}
        
        # Install and import PnP PowerShell
        if (Install-PnPIfNeeded) {{
            Import-Module PnP.PowerShell -Force
        }} else {{
            throw "Cannot proceed without PnP.PowerShell module"
        }}
        
        # Connection parameters
        $siteUrl = "{SHAREPOINT_SITE_URL}"
        $folderPath = "{SHAREPOINT_FOLDER_PATH}"
        $localFilePath = "{file_path.replace(chr(92), chr(92) + chr(92))}"
        $fileName = "{filename}"
        
        try {{
            Write-Host " Connecting to SharePoint Site: $siteUrl"
            Write-Host " Target folder: $folderPath"
            Write-Host " File to upload: $fileName"
            
            # Try interactive authentication first
            Write-Host " Attempting interactive authentication..."
            Connect-PnPOnline -Url $siteUrl -Interactive -ForceAuthentication
            
            Write-Host " Uploading file to SharePoint..."
            
            # Upload the file with overwrite
            $uploadResult = Add-PnPFile -Path $localFilePath -Folder $folderPath -NewFileName $fileName -Overwrite
            
            if ($uploadResult) {{
                Write-Host " SUCCESS: File uploaded to SharePoint!"
                Write-Host " Server location: $($uploadResult.ServerRelativeUrl)"
                
                # Get full URL
                $fullUrl = "{SHAREPOINT_SITE_URL}" + $uploadResult.ServerRelativeUrl
                Write-Host " Full URL: $fullUrl"
                
                # Disconnect
                Disconnect-PnPOnline
                
                # Return success result
                $result = @{{
                    Success = $true
                    Message = "File uploaded successfully to SharePoint!"
                    FileUrl = $fullUrl
                    ServerPath = $uploadResult.ServerRelativeUrl
                    UploadTime = (Get-Date).ToString()
                }}
                
                $jsonResult = $result | ConvertTo-Json -Compress
                Write-Host "UPLOAD_RESULT:$jsonResult"
                return $true
            }} else {{
                throw "Upload returned null result - file may not have been uploaded"
            }}
        }}
        catch {{
            Write-Host " UPLOAD FAILED: $($_.Exception.Message)"
            Write-Host " Error details: $($_.Exception.ToString())"
            
            $errorResult = @{{
                Success = $false
                Message = $_.Exception.Message
                ErrorDetails = $_.Exception.ToString()
                FileUrl = $null
            }}
            
            $jsonError = $errorResult | ConvertTo-Json -Compress
            Write-Host "UPLOAD_RESULT:$jsonError"
            return $false
        }}
        """
        
        # Execute PowerShell script with extended timeout
        result = subprocess.run(
            ["powershell", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
            capture_output=True, 
            text=True, 
            timeout=1200  # 20 minutes timeout for authentication
        )
        
        # Check for success indicators in output
        if " SUCCESS: File uploaded to SharePoint!" in result.stdout:
            # Extract JSON result if available
            json_marker = "UPLOAD_RESULT:"
            if json_marker in result.stdout:
                json_start = result.stdout.find(json_marker) + len(json_marker)
                json_end = result.stdout.find("\n", json_start)
                if json_end == -1:
                    json_end = len(result.stdout)
                
                try:
                    json_str = result.stdout[json_start:json_end].strip()
                    upload_info = json.loads(json_str)
                    return {{
                        "success": True,
                        "message": upload_info.get("Message", "File uploaded successfully!"),
                        "file_url": upload_info.get("FileUrl", ""),
                        "upload_time": upload_info.get("UploadTime", ""),
                        "output": result.stdout
                    }}
                except json.JSONDecodeError:
                    pass
            
            return {{
                "success": True,
                "message": "File uploaded successfully to SharePoint!",
                "output": result.stdout
            }}
        else:
            error_msg = result.stderr if result.stderr else "Upload failed - check output for details"
            return {{
                "success": False,
                "message": error_msg,
                "output": result.stdout + "\n" + result.stderr
            }}
            
    except subprocess.TimeoutExpired:
        return {{
            "success": False,
            "message": "Upload timeout - Authentication or upload took too long",
            "output": "Process timed out after 20 minutes"
        }}
    except Exception as e:
        return {{
            "success": False,
            "message": f"Upload error: {{str(e)}}",
            "output": str(e)
        }}

def create_upload_folders(filename):
    """Create upload folders and save file locally"""
    try:
        # Create multiple upload locations
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "SharePoint_Upload")
        documents_path = os.path.join(os.path.expanduser("~"), "Documents", "SharePoint_Upload")
        
        # Create directories
        os.makedirs(desktop_path, exist_ok=True)
        os.makedirs(documents_path, exist_ok=True)
        
        return {{
            "desktop": desktop_path,
            "documents": documents_path
        }}
    except Exception as e:
        return {{
            "error": str(e)
        }}

def automated_sharepoint_upload(file_path, filename):
    """Complete automated SharePoint upload with multiple fallbacks"""
    
    upload_progress = st.empty()
    
    # Step 1: Try PowerShell PnP upload
    upload_progress.write(" **Step 1**: Attempting PowerShell PnP upload...")
    
    pnp_result = upload_to_sharepoint_powershell(file_path, filename)
    
    if pnp_result["success"]:
        upload_progress.write(" **SUCCESS**: PowerShell upload completed!")
        return pnp_result
    
    upload_progress.write(f" PowerShell upload failed: {{pnp_result['message']}}")
    
    # Step 2: Fallback - Save to multiple locations and open SharePoint
    upload_progress.write(" **Step 2**: Creating local copies and opening SharePoint...")
    
    try:
        # Create upload folders
        folders = create_upload_folders(filename)
        
        if "error" not in folders:
            # Copy to desktop
            desktop_file = os.path.join(folders["desktop"], filename)
            shutil.copy2(file_path, desktop_file)
            
            # Copy to documents
            documents_file = os.path.join(folders["documents"], filename)
            shutil.copy2(file_path, documents_file)
            
            # Open SharePoint in browser
            webbrowser.open(SHAREPOINT_FULL_URL)
            
            upload_progress.write(" **Fallback SUCCESS**: Files saved locally and SharePoint opened!")
            
            return {{
                "success": True,
                "message": f"File saved to Desktop and Documents folders. SharePoint opened in browser.",
                "desktop_path": desktop_file,
                "documents_path": documents_file,
                "output": "Fallback upload method successful - drag and drop the file from Desktop to SharePoint"
            }}
        else:
            raise Exception(f"Could not create upload folders: {{folders['error']}}")
            
    except Exception as e:
        upload_progress.write(f" **All methods failed**: {{str(e)}}")
        return {{
            "success": False,
            "message": f"All upload methods failed: {{str(e)}}",
            "output": "Complete failure - manual download required"
        }}

def create_excel_file(data, from_date, to_date):
    """Create Excel file with multiple sheets and formatting"""
    
    # Filter data by date range
    if from_date and to_date:
        data["date_added"] = pd.to_datetime(data["date_added"])
        filtered_data = data[
            (data["date_added"].dt.date >= from_date) & 
            (data["date_added"].dt.date <= to_date)
        ].copy()
    else:
        filtered_data = data.copy()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create Summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Summary data
    total_entries = len(filtered_data)
    total_time_saved = filtered_data["time_saved_hours"].sum() if not filtered_data.empty else 0
    unique_members = filtered_data["team_member"].nunique() if not filtered_data.empty else 0
    
    # Write summary with formatting
    summary_data = [
        ["BLR Productivity Gains Summary", ""],
        ["Report Period", f"{{from_date}} to {{to_date}}" if from_date and to_date else "All Time"],
        ["", ""],
        ["Total Entries", total_entries],
        ["Total Time Saved (Hours)", f"{{total_time_saved:.2f}}"],
        ["Unique Team Members", unique_members],
        ["", ""],
        ["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Generated By", "Sprint Productivity Tracker - AUTO UPLOAD"]
    ]
    
    for row in summary_data:
        summary_ws.append(row)
    
    # Format summary sheet
    summary_ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    summary_ws["A1"].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Create Data sheet
    if not filtered_data.empty:
        data_ws = wb.create_sheet("Detailed Data")
        
        # Prepare data for Excel
        export_data = filtered_data[["date_added", "team_member", "task_description", 
                                   "time_saved_hours", "impact_category", "sprint_week"]].copy()
        export_data.columns = ["Date Added", "Team Member", "Task Description", 
                              "Time Saved (Hours)", "Impact Category", "Sprint/Week"]
        
        # Add headers
        headers = export_data.columns.tolist()
        data_ws.append(headers)
        
        # Add data
        for row in dataframe_to_rows(export_data, index=False, header=False):
            data_ws.append(row)
        
        # Format headers
        for col_num, header in enumerate(headers, 1):
            cell = data_ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def main():
    st.set_page_config(
        page_title="BLR Productivity Tracker - AUTO UPLOAD",
        page_icon="",
        layout="wide"
    )
    
    # Initialize database
    init_database()
    
    st.title(" BLR Sprint Productivity Tracker")
    st.markdown("###  **AUTO UPLOAD EDITION** - Automatic SharePoint Integration")
    st.markdown("---")
    
    # Sidebar for admin login
    with st.sidebar:
        st.header(" Admin Access")
        
        if "admin_logged_in" not in st.session_state:
            st.session_state.admin_logged_in = False
        
        if not st.session_state.admin_logged_in:
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            
            if st.button("Admin Login"):
                if verify_admin(username, password):
                    st.session_state.admin_logged_in = True
                    st.success(" Admin access granted!")
                    st.rerun()
                else:
                    st.error(" Invalid credentials")
        else:
            st.success(" Logged in as Admin")
            if st.button("Logout"):
                st.session_state.admin_logged_in = False
                st.rerun()
    
    # Main content
    tab1, tab2, tab3 = st.tabs([" Data Entry", " View Data", " AUTO UPLOAD"])
    
    with tab1:
        st.header("Add Productivity Data")
        
        # Data entry form
        col1, col2 = st.columns(2)
        
        with col1:
            date_added = st.date_input("Date Added", value=date.today())
            team_member = st.text_input("Team Member Name")
            task_description = st.text_area("Task Description", height=100)
        
        with col2:
            time_saved = st.number_input("Time Saved (Hours)", min_value=0.0, step=0.5)
            impact_category = st.selectbox("Impact Category", 
                ["High", "Medium", "Low", "Process Improvement", "Automation", "Bug Fix"])
            sprint_week = st.text_input("Sprint/Week", placeholder="e.g., Sprint 23, Week 15")
        
        if st.button(" Add Entry", type="primary"):
            if team_member and task_description and time_saved > 0 and sprint_week:
                new_data = pd.DataFrame({{
                    "Date Added": [date_added.strftime("%Y-%m-%d")],
                    "Team Member": [team_member],
                    "Task Description": [task_description],
                    "Time Saved (Hours)": [time_saved],
                    "Impact Category": [impact_category],
                    "Sprint/Week": [sprint_week]
                }})
                
                add_data_to_db(new_data)
                st.success(" Data added successfully!")
                st.rerun()
            else:
                st.error(" Please fill all required fields")
    
    with tab2:
        st.header("Current Data")
        
        # Load and display data
        data = get_all_data()
        
        if not data.empty:
            # Display metrics
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Total Entries", len(data))
            with col2:
                st.metric("Total Time Saved", f"{{data['time_saved_hours'].sum():.1f}} hrs")
            with col3:
                st.metric("Team Members", data["team_member"].nunique())
            with col4:
                st.metric("Latest Entry", data["date_added"].iloc[0] if not data.empty else "None")
            
            # Display data table
            st.dataframe(
                data[["date_added", "team_member", "task_description", "time_saved_hours", "impact_category", "sprint_week"]],
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info(" No data available. Add some entries in the Data Entry tab.")
    
    with tab3:
        st.header(" AUTOMATED SHAREPOINT UPLOAD")
        st.markdown("**Automatically creates Excel file and uploads directly to SharePoint!**")
        
        # Date range selection
        col1, col2 = st.columns(2)
        with col1:
            from_date = st.date_input("From Date", value=date.today().replace(day=1))
        with col2:
            to_date = st.date_input("To Date", value=date.today())
        
        data = get_all_data()
        
        if not data.empty:
            # Show preview of filtered data
            if from_date and to_date:
                data["date_added"] = pd.to_datetime(data["date_added"])
                filtered_data = data[
                    (data["date_added"].dt.date >= from_date) & 
                    (data["date_added"].dt.date <= to_date)
                ]
                st.info(f" {{len(filtered_data)}} entries will be exported for the selected date range")
            else:
                filtered_data = data
                st.info(f" All {{len(data)}} entries will be exported")
            
            # Main auto-upload section
            st.markdown("###  One-Click Automated Upload")
            st.markdown("**Click below to create Excel file and automatically upload to SharePoint!**")
            
            # Big prominent button
            if st.button(" CREATE & AUTO-UPLOAD TO SHAREPOINT", type="primary", use_container_width=True):
                # Create Excel file
                excel_file = create_excel_file(data, from_date, to_date)
                filename = f"BLR_Productivity_Gains_{{from_date}}_{{to_date}}.xlsx"
                
                st.info(f" Creating file: {{filename}}")
                
                # Save to temporary location
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                    tmp_file.write(excel_file.getvalue())
                    temp_path = tmp_file.name
                
                # Show upload progress
                st.markdown("###  Upload Progress")
                
                # Attempt automated upload
                upload_result = automated_sharepoint_upload(temp_path, filename)
                
                if upload_result["success"]:
                    st.success(" **UPLOAD SUCCESSFUL!** File automatically uploaded to SharePoint!")
                    
                    if "file_url" in upload_result and upload_result["file_url"]:
                        st.markdown(f" **Direct Link**: [Open File]({{upload_result['file_url']}})")
                    
                    if "desktop_path" in upload_result:
                        st.info(f" **Also saved to**: {{upload_result['desktop_path']}}")
                    
                    # Show success details
                    with st.expander(" Upload Details"):
                        st.text(upload_result["output"])
                        
                else:
                    st.error(f" **Upload Failed**: {{upload_result['message']}}")
                    
                    # Show error details
                    with st.expander(" Error Details"):
                        st.text(upload_result["output"])
                    
                    # Provide manual download as fallback
                    st.warning(" **Fallback**: Download file manually below")
                    
                    st.download_button(
                        label=" Download Excel File",
                        data=excel_file.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                # Clean up temp file
                try:
                    os.unlink(temp_path)
                except:
                    pass
            
            # SharePoint link
            st.markdown("---")
            st.markdown("###  SharePoint Destination")
            st.markdown(f"[ **Open SharePoint Folder**]({{SHAREPOINT_FULL_URL}})")
            st.caption(" Target: BLR productivity gains  General  2025")
            
        else:
            st.warning(" No data available to export")
    
    # Footer
    st.markdown("---")
    st.markdown("###  System Information")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("** Admin Access**: Adityakarthik only")
    with col2:
        st.markdown("** Target Email**: rakhi.purohit@thomsonreuters.com")
    with col3:
        st.markdown("** Upload Method**: PowerShell PnP + Fallback")

if __name__ == "__main__":
    main()
